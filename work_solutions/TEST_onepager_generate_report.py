from openpyxl.styles import PatternFill, Font, Side, Border, numbers, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from io import BytesIO
from typing import Any
import sqlalchemy as sqla
import traceback as trb
import datetime as dt
import urllib.parse
import pandas as pd
import openpyxl
import time

"""
One Pager report generation
Each report has multiples sheets
2 types of reports (OnePager, Reject)
Insert multiple tables into single excel sheet
- Different sheets will have different tables

Client
-------
SID = Shred-It
RWCS = Stericycle

Multiple select statements -- will result in their own dataframe

Use functions for styling as styling will most likely be done multiple times
---------------------------------
ACTIVE REJECTCLOSE CANADA :: 6 Sheets
---------------------------------
Summary == 4 tables (8pt font, bold rows, outer borders only)
Loader Rejects PIF Close == 1 table (default font, no borders)
Treatment Complete == 1 table (default font, no borders)
Active Under RWCS == 1 table (default font, no borders)
Cont Writeoff not in Aging == 1 table (default font, no borders)
Active == 1 table (default font, no borders)
---------------------------------
---------------------------------
ACTIVE REJECTCLOSE STERICYCLE US :: 6 Sheets
---------------------------------
Summary == 4 tables (defaultfont, bold rows, outer borders only)
Loader Rejects PIF CLOSE == 1 table (default font, no borders)
Treatment Complete == 1 table (default font, no borders)
Active Under SID == Empty
Cont Writeoff not in Aging == 1 table (default font, no borders)
Active == 1 table (default font, no borders)
---------------------------------
---------------------------------
ACTIVE REJECTCLOSE US :: 6 Sheets
---------------------------------
Summary == 4 tables (8pt font, bold rows, outer borders only)
Loader Rejects PIF Close == 1 table (default font, no borders)
Treatment Complete == 1 table (default font, no borders)
Active Under RWCS == Empty
Cont Writeoff not in Aging == 1 table (default font, no borders)
Active == 1 table (default font, no borders)
---------------------------------
---------------------------------
SID ONE PAGER CANADA :: 2 Sheets
---------------------------------
Placements by Month == 5 tables, 2 images (colored headers [blue/green] + white font, default font, bold rows)
Placements & Performance by Cat == 3 tables (colored headers [blue/green] + white font, default font, bold rows)
---------------------------------
---------------------------------
SID ONE PAGER US :: 2 Sheets
---------------------------------
Placements by Month == 5 tables, 2 images (colored headers [blue/green] + white font, default font, bold rows)
Placements & Performance by Cat == 3 tables (colored headers [blue/green] + white font, default font, bold rows)
---------------------------------
---------------------------------
SID ONE PAGER USD_SUMMARY :: 2 Sheets
---------------------------------
Shows conversion of CAD to USD
Placements by Month == 5 tables, 2 images (colored headers [blue/green] + white font, default font, bold rows)
Placements & Performance by Cat == 3 tables (colored headers [blue/green] + white font, default font, bold rows)
"""

# CONSTANTS FOR COLUMNS AND SHEETS
# Sheets
ACTIVE_SHEETS = ["Summary", "Loader Rejects PIF Close", "Treatment Complete", "Active Under RWCS",
                 "Active Under SID", "Cont Writeoff not in Aging", "Active"]
ONE_PAGER_SHEETS = ["Placements by Month", "Placements & Performance by Cat"]

# Bold these rows
ROWS_TO_BOLD = ["Summary", "Total", "Treatment", "Collection", "Grand Total",
                "Sub Total of Collection Phases", "Difference", "Canadian Total", "US Total", "Combined Total"]
# Apply $ to values in these columns
DOLLARS_COL = ["Original Assigned", "Additional Charges", "Total Assigned", "Paid", "Credit",
               "Stop Amt", "Resolved", "$Exhausted", "$Current", "At ARM Active",
               "Total Assigned to ARM - Current & Active $", "Reconciliation", "Dollars", "ARM",
               "SID Aging", "Difference", "AR Total Balance", "Total Open AR", "AmtDueInARM",
               "DueAmtInClientAging", "Total Past Due", "Principal Received", "PastDueinAging",
               "Referral Amt", "Principal Due"]
# Apply % to values in these columns
PERC_COL = ["Paid %", "Credit %", "Stop Amt %", "Resolved %", "Exhausted %", "Resolved & $Exhuasted %",
            "Active %"]

# Color these headers in OnePager reports
COLOR_OP_SHEET_1 = [5, 12, 14, 15, 16, 21, 22, 23]
COLOR_OP_SHEET_2 = [9, 16, 18, 19, 20, 25, 26, 27]

DROP_SORT_COL = ["Sort", "sort"]

today_date = dt.datetime.today().strftime("%Y.%m.%d")


def sql_connection() -> str:  # establish connection to sql server

    sql_server = "armsql"
    sql_database = "CSS_Local"
    driver = "ODBC Driver 17 for SQL Server"
    connect_setup = f'DRIVER={driver};SERVER={sql_server};DATABASE={sql_database};Trusted_Connection=yes;'

    connect_param = urllib.parse.quote_plus(connect_setup)
    connect_to = f"mssql+pyodbc://@{sql_server}/{sql_database}?driver={driver}&odbc_connect={connect_param}"

    return connect_to


def color_headers(excel_sheet: list[Any], sheet_num: list[int], columns: list[list], font_size: list[int] = None, start_row=1):
    # Color headers blue, then specific column headers green (based on list) within an excel worksheet
    blue_fill = PatternFill(start_color="4C68A2", end_color="4C68A2", fill_type="solid")
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    for idx, sheet in enumerate(excel_sheet):
        if font_size is None:
            font_num = 11
        else:
            font_num = font_size[idx]
        for row in sheet.iter_rows(min_row=start_row, max_row=start_row):
            for cell in row:
                cell.fill = blue_fill
                cell.font = Font(bold=True, color="FFFFFF", size=font_num)

        for col in columns[idx]:
            excel_sheet[idx].cell(row=start_row, column=col).fill = green_fill

    return


# adjust column width for sheet
def adjust_col_width(excel_sheet: list[Any], width_modifier: list[float] = 0.0):
    for idx, sheet in enumerate(excel_sheet):
        for col in sheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # get letter of each column

            for cell in col:  # for each cell in columns
                try:
                    if len(str(cell.value)) > max_length:  # if length of cell greater than last max_length
                        max_length = len(str(cell.value))  # max_length == current length of cell
                except:
                    pass

                # store max_length value and add it with width_modifier
                adjusted_width = (max_length + width_modifier[idx])
                # apply width adjustment to column header cells
                sheet.column_dimensions[column].width = adjusted_width

    return


def bold_rows(excel_sheet: list[Any], bold_values: list[list], font_size: list[int] = None, start_row=0):
    # Bold font in all rows containing values from list (bold_values)
    for idx, sheet in enumerate(excel_sheet):
        if font_size is None:
            bold_font = Font(bold=True, size=11)
        else:
            bold_font = Font(bold=True, size=font_size[idx])
        for row in excel_sheet[idx].iter_rows(min_row=start_row):
            if idx < len(bold_values):
                if row[0].value in bold_values[idx]:
                    for cell in row:
                        cell.font = bold_font
            elif row[0].value in bold_values[idx - 1]:
                for cell in row:
                    cell.font = bold_font

    return


# bold all values in dollars column
def bold_column_dollars(excel_sheet: Any, font_size=11):
    bold_font = Font(bold=True, size=font_size)

    # if cell value == "Dollars", store position
    for row in excel_sheet.iter_rows():
        for cell in row:
            if cell.value == "Dollars":
                cell_num = cell.row

                # for cells starting at position cell_num and ending at max_row
                for cells_below in range(cell_num + 1, excel_sheet.max_row + 1):
                    # bold all text in column B starting at position cell_num
                    excel_sheet['B' + str(cells_below)].font = bold_font

    return


def font_size8(excel_sheet: Any):
    # Change font size of entire worksheet to 8pts
    font_size = Font(size=8)

    for row in excel_sheet.iter_rows():
        for cell in row:
            cell.font = font_size
    return


def find_tables(excel_sheet):  # Find position of all tables
    tables = []  # stores each tables start and end row

    # get non empty rows and store into list
    non_empty_rows = [row for row in excel_sheet.iter_rows() if any(cell.value is not None for cell in row)]

    if not non_empty_rows:  # if empty
        return tables

    # first block/table with the row number of the first non-empty row
    current_block = (non_empty_rows[0][0].row, non_empty_rows[0][-1].row)

    # iterate through each non-empty row to determine where tables start and end
    for row in non_empty_rows:
        # If there is a gap of more than one empty row, set as table cutoff and append
        if row[0].row > current_block[1] + 1:
            tables.append(current_block)
            current_block = (row[0].row, row[0].row)  # start new block/table with current row

        current_block = (current_block[0], row[0].row)  # update end row of current block/table

    tables.append(current_block)  # append last table to list
    return tables


def outer_borders(excel_sheet, font_size=11):
    tables = find_tables(excel_sheet)

    thin_b = Side(style='thin')
    bold_headers = Font(bold=True, size=font_size)

    border = Border(left=thin_b,
                    right=thin_b,
                    top=thin_b,
                    bottom=thin_b)
    # get min and max columns that have data within each table
    for start_row, end_row in tables:
        min_col = min((cell.column for row in excel_sheet.iter_rows(min_row=start_row, max_row=end_row)
                       for cell in row if cell.value is not None))
        max_col = max((cell.column for row in excel_sheet.iter_rows(min_row=start_row, max_row=end_row)
                       for cell in row if cell.value is not None))

        # apply borders to edges of each table
        for row in excel_sheet.iter_rows(min_row=start_row, max_row=end_row,
                                         min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = Border(top=border.top if cell.row == start_row else None,
                                     bottom=border.bottom if cell.row == end_row else None,
                                     left=border.left if cell.column == min_col else None,
                                     right=border.right if cell.column == max_col else None)
                if cell.row == start_row:
                    cell.font = bold_headers


def add_border(df_to_style):
    # Add border to all cells of a dataframe
    df_to_style = df_to_style.style.set_properties(**{'border': '1px solid black'})

    return df_to_style


def cell_to_percent_currency(excel_sheet: list[Any]):
    # Convert cells to Currency
    for sheet in excel_sheet:  # iterate through list of sheets
        for row in sheet.iter_rows():  # iterate through rows in sheet
            for cell in row:
                if isinstance(cell.value, str) and '$' in cell.value:  # if cell == $ convert to float
                    dollar_cell = cell.value.replace('$', '').replace(',', '')
                    try:
                        cell.value = float(dollar_cell)
                        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE  # change cell format
                    except ValueError:
                        pass

    # Convert cells to Percentage
    for sheet in excel_sheet:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.endswith('%'):  # if cell == % convert to float
                    percent_cell = cell.value.replace('%', '')
                    try:
                        cell.value = float(percent_cell)
                        cell.number_format = numbers.FORMAT_PERCENTAGE
                    except ValueError:
                        pass

    return


# align text to Center in column
def center_align_start_date(excel_sheet: Any, start_row=2):
    align_col = 'B'
    end_row = None

    # start at position and end at max length of column (last row that had data in that column)
    for row in range(start_row, len(excel_sheet[align_col]) + 1):
        if excel_sheet[align_col + str(row)].value == 'Accounts':
            end_row = row - 2  # 'Accounts' acts as the end marker (next table)
            break

    # center all text from starting row to ending row
    for row in range(start_row, end_row + 1):
        excel_sheet[align_col + str(row)].alignment = Alignment(horizontal='center')

    return


# Insert images into excel sheet
def insert_image(excel_sheet: Any, start_row=2, column='', picture=''):
    end_row = None

    # start at position and end at max length of column (last row that had data in that column)
    for row in range(start_row, len(excel_sheet[column]) + 1):
        if excel_sheet[column + str(row)].value is None:
            end_row = row + 1
            break
        else:  # if start row is already the end of a column, take that position and add 3 spaces
            end_row = start_row + 3

    # insert image at end row (+1 or 3 spaces down)
    png = picture
    my_ping = openpyxl.drawing.image.Image(png)
    excel_sheet.add_image(my_ping, column + str(end_row))


def remove_format(df):
    # Remove default bolding and borders from dataframe being pushed into Excel

    return df.T.reset_index().T


def insert_dataframes(df_list: list[pd.DataFrame], excel_book: Any, excel_sheet: str,
                      start_row=0, row_spacing=2, option_header=True, option_index=False) -> None:
    # Insert list of dataframes into specific workbook and sheet at position and spacing based on parameter
    for dataframe in df_list:
        dataframe.to_excel(excel_book, startrow=start_row, sheet_name=excel_sheet, index=option_index, header=option_header)
        start_row += len(dataframe.index) + row_spacing
    return


# Save and write data to excel sheet in-memory
def write_excel(excel_name, df_list: list[pd.DataFrame] = pd.DataFrame, sheet_name: str = "", start_row=0,
                row_spacing=2, option_header=True, option_index=False):
    dataframes = df_list
    sheet = sheet_name

    try:
        with pd.ExcelWriter(excel_name, engine="openpyxl", mode="a", if_sheet_exists="overlay") as write_stream:
            for df in dataframes:
                df.to_excel(write_stream, sheet_name=sheet, startrow=start_row, index=option_index, header=option_header)
                start_row += len(df.index) + row_spacing
    except:
        with pd.ExcelWriter(excel_name, engine="openpyxl") as write_stream:
            for df in dataframes:
                df.to_excel(write_stream, sheet_name=sheet, startrow=start_row, index=option_index, header=option_header)
                start_row += len(df.index) + row_spacing

    # reset byte stream pointer to starting position
    excel_name.seek(0)

    return excel_name


def create_onepager_report(sql_query: list[str], name: str):
    start_time = time.time()

    # sql queries -> store data into dataframes list

    # initialize list to store dataframes
    dataframe_list = []

    # establish connection to sql server and execute queries
    with engine.connect() as conn:
        for query in sql_query:
            # pull query data into dataframe variable
            dataframe = pd.read_sql(query, conn)

            # Drop "Sort" column from all dataframes
            for col in DROP_SORT_COL:
                dataframe = dataframe.drop(columns=col, errors='ignore')

            # store dataframe into list
            dataframe_list.append(dataframe)

    # Apply $ or % to columns in all dataframes
    for idx, df in enumerate(dataframe_list):
        for col in df:
            if col in DOLLARS_COL:
                df[col] = df[col].map("${:,}".format)
            if col in PERC_COL:
                df[col] = df[col].map("{:,}%".format)

    # initialize byte-stream like object
    excel_bytes = BytesIO()

    # write to excel sheet in-memory
    if len(dataframe_list) > 6:
        # add borders to all cells in all dataframes
        for idx, df in enumerate(dataframe_list[0::]):
            dataframe_list[idx] = add_border(df)

        excel_book = write_excel(excel_bytes, df_list=dataframe_list[0:5], sheet_name="Placements by Month")
        excel_book = write_excel(excel_bytes, df_list=dataframe_list[5::],
                                 sheet_name="Placements & Performance by Cat")

        # load excel workbook
        book = load_workbook(excel_book)

        # assign excel sheets to variables
        sheet_1 = book["Placements by Month"]
        sheet_2 = book["Placements & Performance by Cat"]

        # adjust col width
        adjust_col_width(excel_sheet=[sheet_1, sheet_2], width_modifier=[4.25, 1.25])

        # convert cells to currency or percent data type in Excel
        cell_to_percent_currency(excel_sheet=[sheet_1, sheet_2])

        # Color headers
        color_headers(excel_sheet=[sheet_1, sheet_2], sheet_num=[1, 1], columns=[COLOR_OP_SHEET_1, COLOR_OP_SHEET_2])

        # Bold specific rows
        bold_rows(excel_sheet=[sheet_1, sheet_2], bold_values=[ROWS_TO_BOLD])

        # Bold all values in Dollars column
        bold_column_dollars(excel_sheet=sheet_2)

        # Center align text in column activity start date
        center_align_start_date(excel_sheet=sheet_2)

        # Insert images into excel sheet
        insert_image(sheet_1, 2, 'H', 'picture_leg.png')
        insert_image(sheet_1, len(sheet_1['A']), 'A', 'picture_nonleg.png')

        # initialize new byte-stream
        temp_book = BytesIO()

        # freeze row in excel
        sheet_2.freeze_panes = sheet_2["A2"]

        # save all changes into new byte-stream
        book.save(temp_book)

        # get data that was just saved
        book1_data = temp_book.getvalue()

        with open(f"SID {today_date} {name}.xlsx", 'wb') as f:
            f.write(book1_data)

        temp_book.seek(0)
        temp_book.truncate(0)
    else:
        dataframe_list[0] = dataframe_list[0].T

        # add borders to all cells in all dataframes
        for idx, df in enumerate(dataframe_list):
            dataframe_list[idx] = add_border(df)

        excel_book = write_excel(excel_bytes, df_list=[dataframe_list[0]], sheet_name="Placements by Month", option_header=False, option_index=True)

        excel_book = write_excel(excel_bytes, df_list=dataframe_list[1::], sheet_name="Placements by Month", start_row=1)
        # load excel workbook
        book = load_workbook(excel_book)

        # assign excel sheets to variables
        sheet_1 = book["Placements by Month"]

        # adjust col width
        adjust_col_width(excel_sheet=[sheet_1], width_modifier=[4.25])

        # convert cells to currency or percent data type in Excel
        cell_to_percent_currency(excel_sheet=[sheet_1])

        # Color headers
        color_headers(excel_sheet=[sheet_1], sheet_num=[1], columns=[COLOR_OP_SHEET_1], start_row=2)

        # Bold specific rows
        bold_rows(excel_sheet=[sheet_1], bold_values=[ROWS_TO_BOLD], start_row=2)


        # Insert images into excel sheet
        insert_image(sheet_1, 2, 'G', 'picture_leg.png')
        insert_image(sheet_1, len(sheet_1['A']), 'A', 'picture_nonleg.png')

        # initialize new byte-stream
        temp_book = BytesIO()

        # save all changes into new byte-stream
        book.save(temp_book)

        # get data that was just saved
        book1_data = temp_book.getvalue()

        with open(f"SID {today_date} {name}.xlsx", 'wb') as f:
            f.write(book1_data)

        temp_book.seek(0)
        temp_book.truncate(0)


    print(f"{name} took {time.time() - start_time} seconds")

    return


def create_rejectclose_report(sql_query: list[str], name: str, option_fontsize=True):
    start_time = time.time()

    dataframe_list = []

    with engine.connect() as conn:
        for q in sql_query:
            dataframe = pd.read_sql(q, conn)

            for col in DROP_SORT_COL:
                dataframe = dataframe.drop(columns=col, errors='ignore')

            dataframe_list.append(dataframe)

    for df in dataframe_list[0:6]:
        for col in df:
            if col in DOLLARS_COL:
                # df[col] = df[col].map("${:,}".format)
                df[col] = df[col].apply(lambda x: "${:,.2f}".format(float(x)))
            if col in PERC_COL:
                df[col] = df[col].map("{:,}%".format)

    for idx, df in enumerate(dataframe_list[0::]):
        dataframe_list[idx] = remove_format(df)

    excel_bytes = BytesIO()

    excel_book = write_excel(excel_bytes, df_list=dataframe_list[0:4], sheet_name="Summary", option_header=False)
    excel_book = write_excel(excel_bytes, df_list=[dataframe_list[4]], sheet_name="Loader Rejects PIF CLOSE",
                             option_header=False)
    excel_book = write_excel(excel_bytes, df_list=[dataframe_list[5]], sheet_name="Treatment Completed",
                             option_header=False)
    excel_book = write_excel(excel_bytes, df_list=[dataframe_list[6]], sheet_name="Active Under RWCS",
                             option_header=False)
    excel_book = write_excel(excel_bytes, df_list=[dataframe_list[7]], sheet_name="Cont Writeoff not in Aging",
                             option_header=False)
    excel_book = write_excel(excel_bytes, df_list=[dataframe_list[8]], sheet_name="Active",
                             option_header=False)

    book = load_workbook(excel_book)
    sheet_1 = book["Summary"]
    sheet_2 = book["Loader Rejects PIF CLOSE"]
    sheet_3 = book["Treatment Completed"]
    sheet_4 = book["Active Under RWCS"]
    sheet_5 = book["Cont Writeoff not in Aging"]
    sheet_6 = book["Active"]
    sheet_list = [sheet_1, sheet_2, sheet_3, sheet_4, sheet_5, sheet_6]
    if option_fontsize:
        font_size8(excel_sheet=sheet_1)
        font_sizenum = 8
    else:
        font_sizenum = 11
    adjust_col_width(excel_sheet=sheet_list[0:5], width_modifier=[3, 3, 1, 1, 1])
    cell_to_percent_currency(excel_sheet=sheet_list[0:4])
    bold_rows(excel_sheet=[sheet_1], bold_values=[ROWS_TO_BOLD], font_size=[font_sizenum])
    outer_borders(sheet_1, font_size=font_sizenum)

    temp_book = BytesIO()
    book.save(temp_book)

    book1_data = temp_book.getvalue()
    with open(f"{name}_{today_date}.xlsx", 'wb') as f:
        f.write(book1_data)

    temp_book.seek(0)
    temp_book.truncate(0)
    print(f"{name} took {time.time() - start_time} seconds")


def main():
    op_ca_query = ["select * from dbo.tmp_OP_SID_OPCA_T1_S1",
                   "select * from dbo.tmp_OP_SID_OPCA_T1_S2",
                   "select * from dbo.tmp_OP_SID_OPCA_T1_S3",
                   "select * from dbo.tmp_OP_SID_OPCA_T1_S4",
                   "select * from dbo.tmp_OP_SID_OPCA_T1_S5",
                   "select * from dbo.tmp_OP_SID_OPCA_T2_S1",
                   "select * from dbo.tmp_OP_SID_OPCA_T2_S2",
                   "select * from dbo.tmp_OP_SID_OPCA_T2_S3"]

    op_us_query = ["select * from dbo.tmp_OP_SID_OPUS_T1_S1",
                   "select * from dbo.tmp_OP_SID_OPUS_T1_S2",
                   "select * from dbo.tmp_OP_SID_OPUS_T1_S3",
                   "select * from dbo.tmp_OP_SID_OPUS_T1_S4",
                   "select * from dbo.tmp_OP_SID_OPUS_T1_S5",
                   "select * from dbo.tmp_OP_SID_OPUS_T2_S1",
                   "select * from dbo.tmp_OP_SID_OPUS_T2_S2",
                   "select * from dbo.tmp_OP_SID_OPUS_T2_S3"]

    op_usca_query = ["SELECT * FROM tmp_OP_SID_OPUSCA_T1_S1",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S2",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S3",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S4",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S5",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S6"]

    op_arca_query = ["SELECT * FROM dbo.tmp_OP_SID_ARCA_T1_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T1_S2",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T1_S3",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T1_S4",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T2_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T3_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T4_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T5_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T6_S1"]

    op_rwcs_query = ["SELECT * FROM tmp_OP_RWCS_ARUS_T1_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T1_S2",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T1_S3",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T1_S4",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T2_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T3_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T4_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T5_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T6_S1"]

    op_arus_query = ["SELECT * FROM dbo.tmp_OP_SID_ARUS_T1_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T1_S2",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T1_S3",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T1_S4",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T2_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T3_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T4_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T5_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T6_S1"]



    try:
        start_time = time.time()
        create_onepager_report(op_ca_query, "TEST_One Pager_Canada")

        create_onepager_report(op_us_query, "TEST_One Pager_US")

        create_onepager_report(op_usca_query, "TEST_One Pager_USD_Summary")

        create_rejectclose_report(op_arca_query, "TEST_Active_RejectClose_Canada")

        create_rejectclose_report(op_rwcs_query, "TEST_Active_RejectClose_Stericycle_US", option_fontsize=False)

        create_rejectclose_report(op_arus_query, "TEST_Active_RejectClose_US")

        print("\nTotal execution took %s seconds" % (time.time() - start_time))
    except Exception as e:
        traceback_error = trb.format_exc()
        print(f"An error occurred: {e}\n{traceback_error}")


# Establish SQL Connection
engine = sqla.create_engine(sql_connection())
main()

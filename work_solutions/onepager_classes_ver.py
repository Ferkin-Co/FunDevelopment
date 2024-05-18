from openpyxl.styles import PatternFill, Font, Side, Border, numbers, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import sqlalchemy as sqla
import traceback as trb
from typing import Any
from io import BytesIO
import datetime as dt
import urllib.parse
import pandas as pd
import openpyxl

# CONSTANTS FOR COLUMNS AND SHEETS
# Sheets
ACTIVE_SHEETS = ["Summary", "Loader Rejects PIF Close", "Treatment Complete", "Active Under RWCS",
                 "Active Under SID", "Cont Writeoff not in Aging", "Active"]
ONE_PAGER_SHEETS = ["Placements by Month", "Placements & Performance by Cat"]

# Bold these rows
ROWS_TO_BOLD = ["Summary", "Total", "Treatment", "Collection", "Grand Total",
                "Sub Total of Collection Phases", "Difference", "Canadian Total", "US Total", "Combined Total"]
# Apply $ to values in these columns
DOLLARS_COL = ["Original Assigned", "Additional Charges", "Total Assigned", "Paid", "Credit",
               "Stop Amt", "Resolved", "$Exhausted", "$Current", "At ARM Active",
               "Total Assigned to ARM - Current & Active $", "Reconciliation", "Dollars", "ARM",
               "SID Aging", "Difference", "AR Total Balance", "Total Open AR", "AmtDueInARM",
               "DueAmtInClientAging", "Total Past Due", "Principal Received", "PastDueinAging",
               "Referral Amt", "Principal Due"]
# Apply % to values in these columns
PERC_COL = ["Paid %", "Credit %", "Stop Amt %", "Resolved %", "Exhausted %", "Resolved & $Exhuasted %",
            "Active %"]

# Color these headers in OnePager reports
COLOR_OP_SHEET_1 = [5, 12, 14, 15, 16, 21, 22, 23]
COLOR_OP_SHEET_2 = [9, 16, 18, 19, 20, 25, 26, 27]

DROP_SORT_COL = ["Sort", "sort"]

today_date = dt.datetime.today().strftime("%Y.%m.%d")


def sql_connection() -> str:  # establish connection to sql server

    sql_server = "armsql"
    sql_database = "CSS_Local"
    driver = "ODBC Driver 17 for SQL Server"
    connect_setup = f'DRIVER={driver};SERVER={sql_server};DATABASE={sql_database};Trusted_Connection=yes;'

    connect_param = urllib.parse.quote_plus(connect_setup)
    connect_to = f"mssql+pyodbc://@{sql_server}/{sql_database}?driver={driver}&odbc_connect={connect_param}"

    return connect_to


class ExcelReport:
    def __init__(self, workbook_name: str):
        self.workbook_name = workbook_name
        self.book = openpyxl.Workbook()
        self.byte_stream = BytesIO()

    def save(self):
        self.book.save(self.workbook_name)

    def add_sheets(self, df_list: list[pd.DataFrame], sheet_name: str, start_row=0, row_spacing=2, option_header=True, option_index=False):
        with pd.ExcelWriter(self.byte_stream, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            for df in df_list:
                df.to_excel(writer, sheet_name=sheet_name, startrow=start_row,
                            header=option_header, index=option_index)
                start_row += len(df.index) + row_spacing
        self.byte_stream.seek(0)
        self.book = load_workbook(self.byte_stream)

    def write_excel(self, df_list: list[Any], sheet_name: str,
                    start_row=0, row_spacing=2, option_header=True, option_index=False):
        self.add_sheets(df_list, sheet_name, start_row, row_spacing, option_header, option_index)
        self.book.save(self.workbook_name)


class StyleExcel:
    def __init__(self, excel_sheet: Any):
        self.excel_sheet = excel_sheet

    def find_tables(self):
        tables = []

        non_empty_rows = [row for row in self.excel_sheet.iter_rows() if any(cell.value is not None for cell in row)]

        if not non_empty_rows:
            return tables

        current_block = (non_empty_rows[0][0].row, non_empty_rows[0][-1].row)

        for row in non_empty_rows:
            if row[0].row > current_block[1] + 1:
                tables.append(current_block)
                current_block = (row[0].row, row[0].row)

            current_block = (current_block[0], row[0].row)

        tables.append(current_block)
        return tables

    def color_headers(self, columns: list[list], font_size: list[int] = None, start_row=1):
        blue_fill = PatternFill(start_color = "4C68A2", end_color="4C68A2",
                                fill_type="solid")
        green_fill = PatternFill(start_color="00B050", end_color="00B050",
        fill_type="solid")

        for idx, sheet in enumerate(self.excel_sheet):
            font_num = font_size[idx] if font_size else 11
            for row in sheet.iter_rows(min_row=start_row, max_row=start_row):
                for cell in row:
                    cell.fill = blue_fill
                    cell.font = Font(bold=True, color="FFFFFF", size=font_num)
            for col in columns[idx]:
                self.excel_sheet[idx].cell(row=start_row, column=col).fill = green_fill

    def adjust_col_width(self, width_modifier: list[float] = 0.0):
        for idx, sheet in enumerate(self.excel_sheet):
            for col in sheet.columns:
                max_length = 0
                column = get_column_letter(col[0].column)

                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                    adjusted_width = (max_length+width_modifier[idx])
                    sheet.column_dimensions[column].width = adjusted_width

    def bold_rows(self, bold_values: list[list], font_size: list[int] = None, start_row=0):
        for idx, sheet in enumerate(self.excel_sheet):
            bold_font = Font(bold=True, size=font_size[idx] if font_size else 11)
            for row in self.excel_sheet[idx].iter_rows(min_row=start_row):
                if row[0].value in bold_values[idx]:
                    for cell in row:
                        cell.font = bold_font

    def bold_column_dollars(self, font_size=11):
        bold_font = Font(bold=True, size=font_size)

        for row in self.excel_sheet.iter_rows():
            for cell in row:
                if cell.value == "Dollars":
                    cell_num = cell.row

                    for cells_below in range(cell_num +1,
                                             self.excel_sheet.max_row + 1):
                        self.excel_sheet['B' + str(cells_below)].font = bold_font

    def font_size8(self):
        font_size = Font(size=8)

        for row in self.excel_sheet.iter_rows():
            for cell in row:
                cell.font = font_size

    def outer_borders(self, font_size=11):
        tables = self.find_tables()

        thin_b = Side(style='thin')
        bold_headers = Font(bold=True, size=font_size)

        border = Border(left=thin_b,
                        right=thin_b,
                        top=thin_b,
                        bottom=thin_b)

        for start_row, end_row in tables:
            min_col = min((cell.column for row in self.excel_sheet.iter_rows(
                min_row=start_row, max_row=end_row)
                           for cell in row if cell.value is not None))
            max_col = max((cell.column for row in self.excel_sheet.iter_rows(
                min_row=start_row, max_row=end_row)
                           for cell in row if cell.value is not None))

            for row in self.excel_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    cell.border = Border(top=border.top if cell.row == start_row else None,
                                         bottom=border.bottom if cell.row == end_row else None,
                                         left=border.left if cell.column == min_col else None,
                                         right=border.right if cell.column == max_col else None
                                         )
                    if cell.row == start_row:
                        cell.font = bold_headers

    def cell_to_percent_currency(self):
        for sheet in self.excel_sheet:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and '$' in cell.value:
                        dollar_cell = cell.value.replace('$', '').replace(',', '')
                        try:
                            cell.value = float(dollar_cell)
                            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                        except ValueError:
                            pass

        for sheet in self.excel_sheet:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.endswith('%'):
                        percent_cell = cell.value.replace('%', '')
                        try:
                            cell.value = float(percent_cell)
                            cell.number_format = numbers.FORMAT_PERCENTAGE
                        except ValueError:
                            pass

    def center_align_start_date(self, start_row=2):
        align_col = 'B'
        end_row = None

        for row in range(start_row, len(self.excel_sheet[align_col]) + 1):
            if self.excel_sheet[align_col + str(row)].value == 'Accounts':
                end_row = row - 2
                break
        for row in range(start_row, end_row + 1):
            self.excel_sheet[align_col + str(row)].alignment = Alignment(horizontal='center')

    def insert_image(self, start_row=2, column='', picture=''):
        end_row = None

        for row in range(start_row, len(self.excel_sheet[column]) + 1):
            if self.excel_sheet[column + str(row)].value is None:
                end_row = row + 1
                break
            else:
                end_row = start_row + 3

        # png = picture
        # my_png = openpyxl.drawing.image.Image(png)
        # self.excel_sheet.add_image(my_png, column + str(end_row))
        img = Image(picture)
        self.excel_sheet.add_image(img, column + str(end_row))


class GenerateReport:
    def __init__(self, sqlconn_string: str):
        self.engine = sqla.create_engine(sqlconn_string)

    def execute_queries(self, queries: list[str]) -> list[pd.DataFrame]:
        dataframe_list = []
        with self.engine.connect() as conn:
            for query in queries:
                dataframe = pd.read_sql(query, conn)
                for col in DROP_SORT_COL:
                    dataframe = dataframe.drop(columns=col, errors='ignore')
                dataframe_list.append(dataframe)
        return dataframe_list

    def create_onepager_report(self, sql_query: list[str], name: str):
        dataframe_list = self.execute_queries(sql_query)
        dataframe_list = self.map_symbols(dataframe_list)

        excel_report = ExcelReport(f"SID {today_date}_{name}.xlsx")

        if len(dataframe_list) > 6:
            dataframe_list = [self.add_border(df) for df in dataframe_list]
            excel_report.write_excel(dataframe_list[0:5], "Placements by Month")
            excel_report.write_excel(dataframe_list[5:], "Placements & Performance by Cat")

            sheet_1 = excel_report.book["Placements by Month"]
            sheet_2 = excel_report.book["Placements & Performance by Cat"]

            styler = StyleExcel([sheet_1, sheet_2])
            styler.adjust_col_width([4.25, 1.25])
            styler.cell_to_percent_currency()
            styler.color_headers(columns=[COLOR_OP_SHEET_1, COLOR_OP_SHEET_2])
            styler.bold_rows([ROWS_TO_BOLD, ROWS_TO_BOLD])
            styler.bold_column_dollars()
            styler.center_align_start_date(start_row=2)
            styler.insert_image(start_row=2, column='H', picture='picture_leg.png')
            styler.insert_image(start_row=len(sheet_1['A']), column='A',
                                picture='picture_nonleg.png')
            sheet_2.freeze_panes = sheet_2["A2"]
            excel_report.save()
        else:
            dataframe_list[0] = dataframe_list[0].T
            dataframe_list = [self.add_border(df) for df in dataframe_list]
            excel_report.write_excel([dataframe_list[0]], "Placements by Month", option_header=False, option_index=True)
            excel_report.write_excel(dataframe_list[1:], "Placements by Month", start_row=1)

            sheet_1 = excel_report.book["Placements by Month"]

            styler = StyleExcel([sheet_1])
            styler.adjust_col_width([4.25])
            styler.cell_to_percent_currency()
            styler.color_headers([COLOR_OP_SHEET_1], start_row=2)
            styler.bold_rows([ROWS_TO_BOLD], start_row=2)
            styler.insert_image(start_row=2, column='G', picture='picture_leg.png')
            styler.insert_image(start_row=len(sheet_1['A']), column='A', picture='picture_nonleg.png')

            excel_report.save()

    def create_rejectclose_report(self, sql_query: list[str], name:str, option_fontsize=True):
        dataframe_list = self.execute_queries(sql_query)
        dataframe_list = self.map_symbols(dataframe_list)

        excel_report = ExcelReport(f"{name}_{today_date}.xlsx")
        excel_report.write_excel(dataframe_list[0:4], sheet_name="Summary",
                                 option_header=False)
        excel_report.write_excel([dataframe_list[4]], "Loader Rejects PIF CLOSE", option_header=False)
        excel_report.write_excel([dataframe_list[5]], "Treatment Completed", option_header=False)
        excel_report.write_excel([dataframe_list[6]], "Active Under RWCS", option_header=False)
        excel_report.write_excel([dataframe_list[7]], "Cont Writeoff not in Aging", option_header=False)
        excel_report.write_excel([dataframe_list[8]], "Active", option_header=False)

        sheet_1 = excel_report.book["Summary"]
        sheet_2 = excel_report.book["Loader Rejects PIF CLOSE"]
        sheet_3 = excel_report.book["Treatment Completed"]
        sheet_4 = excel_report.book["Active Under RWCS"]
        sheet_5 = excel_report.book["Cont Writeoff not in Aging"]
        sheet_6 = excel_report.book["Active"]

        sheet_list = [sheet_1, sheet_2, sheet_3, sheet_4, sheet_5, sheet_6]

        styler = StyleExcel(sheet_list)
        if option_fontsize:
            styler.font_size8()
            font_sizenum = 8
        else:
            font_sizenum = 11

        styler.adjust_col_width([3, 3, 1, 1, 1])
        styler.cell_to_percent_currency()
        styler.bold_rows(bold_values=[ROWS_TO_BOLD], font_size=[font_sizenum])

        excel_report.save()

    @staticmethod
    def add_border(df: pd.DataFrame):
        return df.style.set_properties(**{'border': '1px solid black'})

    @staticmethod
    def map_symbols(dataframe_list: list[pd.DataFrame]):
        for df in dataframe_list:
            for col in df.columns:
                if col in DOLLARS_COL:
                    df[col] = df[col].map("${:,}".format)
                if col in PERC_COL:
                    df[col] = df[col].map("{:,}%")
        return dataframe_list


def main():
    sql_conn = sql_connection()
    report_generator = GenerateReport(sql_conn)

    op_ca_query = ["select * from dbo.tmp_OP_SID_OPCA_T1_S1",
                   "select * from dbo.tmp_OP_SID_OPCA_T1_S2",
                   "select * from dbo.tmp_OP_SID_OPCA_T1_S3",
                   "select * from dbo.tmp_OP_SID_OPCA_T1_S4",
                   "select * from dbo.tmp_OP_SID_OPCA_T1_S5",
                   "select * from dbo.tmp_OP_SID_OPCA_T2_S1",
                   "select * from dbo.tmp_OP_SID_OPCA_T2_S2",
                   "select * from dbo.tmp_OP_SID_OPCA_T2_S3"]

    op_us_query = ["select * from dbo.tmp_OP_SID_OPUS_T1_S1",
                   "select * from dbo.tmp_OP_SID_OPUS_T1_S2",
                   "select * from dbo.tmp_OP_SID_OPUS_T1_S3",
                   "select * from dbo.tmp_OP_SID_OPUS_T1_S4",
                   "select * from dbo.tmp_OP_SID_OPUS_T1_S5",
                   "select * from dbo.tmp_OP_SID_OPUS_T2_S1",
                   "select * from dbo.tmp_OP_SID_OPUS_T2_S2",
                   "select * from dbo.tmp_OP_SID_OPUS_T2_S3"]

    op_usca_query = ["SELECT * FROM tmp_OP_SID_OPUSCA_T1_S1",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S2",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S3",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S4",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S5",
                     "SELECT * FROM tmp_OP_SID_OPUSCA_T1_S6"]

    op_arca_query = ["SELECT * FROM dbo.tmp_OP_SID_ARCA_T1_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T1_S2",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T1_S3",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T1_S4",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T2_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T3_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T4_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T5_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARCA_T6_S1"]

    op_rwcs_query = ["SELECT * FROM tmp_OP_RWCS_ARUS_T1_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T1_S2",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T1_S3",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T1_S4",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T2_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T3_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T4_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T5_S1",
                     "SELECT * FROM tmp_OP_RWCS_ARUS_T6_S1"]

    op_arus_query = ["SELECT * FROM dbo.tmp_OP_SID_ARUS_T1_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T1_S2",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T1_S3",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T1_S4",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T2_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T3_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T4_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T5_S1",
                     "SELECT * FROM dbo.tmp_OP_SID_ARUS_T6_S1"]

    try:
        report_generator.create_onepager_report(op_ca_query, "TEST_One Pager_Canada")
        report_generator.create_onepager_report(op_us_query, "TEST_One Pager_US")
        report_generator.create_onepager_report(op_usca_query, "TEST_One Pager_USD_Summary")
        report_generator.create_rejectclose_report(op_arca_query, "TEST_Active_RejectClose_Canada")
        report_generator.create_rejectclose_report(op_rwcs_query, "TEST_Active_RejectClose_Stericycle_US", option_fontsize=False)
        report_generator.create_rejectclose_report(op_arus_query, "TEST_Active_RejectClose_US")

    except Exception as e:
        traceback_error = trb.format_exc()
        print(f"An error occurred: {e}\n{traceback_error}")
